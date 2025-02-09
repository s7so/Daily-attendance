from datetime import datetime, timedelta
from typing import Optional

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGroupBox,
    QLabel, QPushButton, QTableWidget, QTableWidgetItem,
    QComboBox, QDateEdit, QTabWidget, QFileDialog, QMessageBox,
    QFrame, QSpinBox
)
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtGui import QColor, QFont, QTextDocument
from PyQt6.QtPrintSupport import QPrinter, QPrintDialog

import pandas as pd
from openpyxl import Workbook

from ..styles import (
    BUTTON_STYLE, TABLE_STYLE, INPUT_STYLE,
    LABEL_STYLE, GROUP_BOX_STYLE,
    FONT_CONFIG, LAYOUT_SPACING, WIDGET_MARGINS
)
from ...database.departments_db import DepartmentsDatabase


class ReportsTab(QWidget):
    def __init__(self, db: DepartmentsDatabase):
        super().__init__()
        self.db = db
        self.setup_ui()
        
    def setup_ui(self):
        # Create main layout
        layout = QVBoxLayout(self)
        layout.setSpacing(LAYOUT_SPACING * 2)
        layout.setContentsMargins(WIDGET_MARGINS * 2, WIDGET_MARGINS * 2, 
                                WIDGET_MARGINS * 2, WIDGET_MARGINS * 2)
        
        # Create attendance report group
        self.setup_attendance_group(layout)
        
        # Create overtime report group
        self.setup_overtime_group(layout)
        
        # Load initial data
        self.load_data()
        
    def setup_attendance_group(self, parent_layout):
        """Setup the attendance report group."""
        attendance_group = QGroupBox("تقرير الحضور اليومي")
        attendance_group.setStyleSheet(GROUP_BOX_STYLE + """
            QGroupBox {
                padding: 20px;
                margin-top: 15px;
                border-radius: 8px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 20px;
                padding: 0 10px;
                color: #495057;
                font-weight: bold;
            }
        """)
        attendance_layout = QVBoxLayout(attendance_group)
        attendance_layout.setSpacing(LAYOUT_SPACING)
        attendance_layout.setContentsMargins(20, 20, 20, 20)
        
        # Attendance filters
        filters_layout = QHBoxLayout()
        
        # Department filter
        dept_label = QLabel("القسم:")
        dept_label.setStyleSheet(LABEL_STYLE + """
            QLabel {
                font-size: 14px;
                font-weight: bold;
                color: #495057;
            }
        """)
        self.dept_combo = QComboBox()
        self.dept_combo.setStyleSheet(INPUT_STYLE + """
            QComboBox {
                padding: 8px 15px;
                border-radius: 6px;
                min-width: 200px;
                background-color: white;
            }
            QComboBox::drop-down {
                border: none;
                padding-right: 10px;
            }
            QComboBox::down-arrow {
                image: none;
                border: none;
            }
            QComboBox:hover {
                border-color: #4dabf7;
            }
        """)
        self.dept_combo.addItem("كل الأقسام", None)
        
        # Date filter
        date_label = QLabel("التاريخ:")
        date_label.setStyleSheet(dept_label.styleSheet())
        self.date_edit = QDateEdit()
        self.date_edit.setStyleSheet(INPUT_STYLE + """
            QDateEdit {
                padding: 8px 15px;
                border-radius: 6px;
                min-width: 150px;
                background-color: white;
            }
            QDateEdit::drop-down {
                border: none;
                padding-right: 10px;
            }
            QDateEdit:hover {
                border-color: #4dabf7;
            }
        """)
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDate(QDate.currentDate())
        
        # Add filters to layout
        filters_layout.addWidget(dept_label)
        filters_layout.addWidget(self.dept_combo)
        filters_layout.addSpacing(20)
        filters_layout.addWidget(date_label)
        filters_layout.addWidget(self.date_edit)
        filters_layout.addStretch()
        
        # Add refresh button
        refresh_btn = QPushButton("تحديث")
        refresh_btn.setStyleSheet(BUTTON_STYLE + """
            QPushButton {
                padding: 8px 20px;
                font-size: 14px;
                font-weight: bold;
            }
        """)
        refresh_btn.clicked.connect(self.refresh_attendance_report)
        filters_layout.addWidget(refresh_btn)
        
        attendance_layout.addLayout(filters_layout)
        
        # Add separator
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setStyleSheet("background-color: #dee2e6;")
        attendance_layout.addWidget(line)
        
        # Attendance table
        self.attendance_table = QTableWidget()
        self.attendance_table.setStyleSheet(TABLE_STYLE + """
            QTableWidget {
                border: 1px solid #dee2e6;
                border-radius: 8px;
                background-color: white;
                gridline-color: #e9ecef;
            }
            QTableWidget::item {
                padding: 8px;
                border-bottom: 1px solid #e9ecef;
            }
            QTableWidget::item:selected {
                background-color: #e7f5ff;
                color: #1864ab;
            }
            QHeaderView::section {
                background-color: #f8f9fa;
                padding: 10px;
                border: none;
                border-bottom: 2px solid #dee2e6;
                font-weight: bold;
                color: #495057;
            }
        """)
        self.setup_attendance_table()
        attendance_layout.addWidget(self.attendance_table)
        
        # Export buttons
        export_layout = QHBoxLayout()
        
        excel_btn = QPushButton("تصدير Excel")
        excel_btn.setStyleSheet(BUTTON_STYLE)
        excel_btn.clicked.connect(lambda: self.export_report("attendance", "excel"))
        
        pdf_btn = QPushButton("تصدير PDF")
        pdf_btn.setStyleSheet(BUTTON_STYLE)
        pdf_btn.clicked.connect(lambda: self.export_report("attendance", "pdf"))
        
        print_btn = QPushButton("طباعة")
        print_btn.setStyleSheet(BUTTON_STYLE)
        print_btn.clicked.connect(lambda: self.print_report("attendance"))
        
        export_layout.addWidget(excel_btn)
        export_layout.addWidget(pdf_btn)
        export_layout.addWidget(print_btn)
        export_layout.addStretch()
        
        attendance_layout.addLayout(export_layout)
        
        parent_layout.addWidget(attendance_group)
        
    def setup_overtime_group(self, parent_layout):
        """Setup the overtime report group."""
        overtime_group = QGroupBox("تقرير الساعات الإضافية")
        overtime_group.setStyleSheet(GROUP_BOX_STYLE + """
            QGroupBox {
                padding: 20px;
                margin-top: 15px;
                border-radius: 8px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 20px;
                padding: 0 10px;
                color: #495057;
                font-weight: bold;
            }
        """)
        overtime_layout = QVBoxLayout(overtime_group)
        overtime_layout.setSpacing(LAYOUT_SPACING)
        overtime_layout.setContentsMargins(20, 20, 20, 20)
        
        # Overtime filters
        filters_layout = QHBoxLayout()
        
        # Department filter
        dept_label = QLabel("القسم:")
        dept_label.setStyleSheet(LABEL_STYLE + """
            QLabel {
                font-size: 14px;
                font-weight: bold;
                color: #495057;
            }
        """)
        self.overtime_dept_combo = QComboBox()
        self.overtime_dept_combo.setStyleSheet(INPUT_STYLE + """
            QComboBox {
                padding: 8px 15px;
                border-radius: 6px;
                min-width: 200px;
                background-color: white;
            }
            QComboBox::drop-down {
                border: none;
                padding-right: 10px;
            }
            QComboBox::down-arrow {
                image: none;
                border: none;
            }
            QComboBox:hover {
                border-color: #4dabf7;
            }
        """)
        self.overtime_dept_combo.addItem("كل الأقسام", None)
        
        # Date range filter
        from_label = QLabel("من:")
        from_label.setStyleSheet(dept_label.styleSheet())
        self.from_date = QDateEdit()
        self.from_date.setStyleSheet(INPUT_STYLE + """
            QDateEdit {
                padding: 8px 15px;
                border-radius: 6px;
                min-width: 150px;
                background-color: white;
            }
            QDateEdit::drop-down {
                border: none;
                padding-right: 10px;
            }
            QDateEdit:hover {
                border-color: #4dabf7;
            }
        """)
        self.from_date.setCalendarPopup(True)
        self.from_date.setDate(QDate.currentDate().addDays(-30))
        
        to_label = QLabel("إلى:")
        to_label.setStyleSheet(dept_label.styleSheet())
        self.to_date = QDateEdit()
        self.to_date.setStyleSheet(self.from_date.styleSheet())
        self.to_date.setCalendarPopup(True)
        self.to_date.setDate(QDate.currentDate())
        
        # Add filters to layout
        filters_layout.addWidget(dept_label)
        filters_layout.addWidget(self.overtime_dept_combo)
        filters_layout.addSpacing(20)
        filters_layout.addWidget(from_label)
        filters_layout.addWidget(self.from_date)
        filters_layout.addSpacing(10)
        filters_layout.addWidget(to_label)
        filters_layout.addWidget(self.to_date)
        filters_layout.addStretch()
        
        # Add refresh button
        refresh_btn = QPushButton("تحديث")
        refresh_btn.setStyleSheet(BUTTON_STYLE + """
            QPushButton {
                padding: 8px 20px;
                font-size: 14px;
                font-weight: bold;
            }
        """)
        refresh_btn.clicked.connect(self.refresh_overtime_report)
        filters_layout.addWidget(refresh_btn)
        
        # إضافة عنصر emp_combo المفقود
        emp_label = QLabel("الموظف:")
        self.emp_combo = QComboBox()
        filters_layout.addWidget(emp_label)
        filters_layout.addWidget(self.emp_combo)
        filters_layout.addStretch()
        
        overtime_layout.addLayout(filters_layout)
        
        # Add separator
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setStyleSheet("background-color: #dee2e6;")
        overtime_layout.addWidget(line)
        
        # Overtime table
        self.overtime_table = QTableWidget()
        self.overtime_table.setStyleSheet(TABLE_STYLE + """
            QTableWidget {
                border: 1px solid #dee2e6;
                border-radius: 8px;
                background-color: white;
                gridline-color: #e9ecef;
            }
            QTableWidget::item {
                padding: 8px;
                border-bottom: 1px solid #e9ecef;
            }
            QTableWidget::item:selected {
                background-color: #e7f5ff;
                color: #1864ab;
            }
            QHeaderView::section {
                background-color: #f8f9fa;
                padding: 10px;
                border: none;
                border-bottom: 2px solid #dee2e6;
                font-weight: bold;
                color: #495057;
            }
        """)
        self.setup_overtime_table()
        overtime_layout.addWidget(self.overtime_table)
        
        # Export buttons
        export_layout = QHBoxLayout()
        
        excel_btn = QPushButton("تصدير Excel")
        excel_btn.setStyleSheet(BUTTON_STYLE)
        excel_btn.clicked.connect(lambda: self.export_report("overtime", "excel"))
        
        pdf_btn = QPushButton("تصدير PDF")
        pdf_btn.setStyleSheet(BUTTON_STYLE)
        pdf_btn.clicked.connect(lambda: self.export_report("overtime", "pdf"))
        
        print_btn = QPushButton("طباعة")
        print_btn.setStyleSheet(BUTTON_STYLE)
        print_btn.clicked.connect(lambda: self.print_report("overtime"))
        
        export_layout.addWidget(excel_btn)
        export_layout.addWidget(pdf_btn)
        export_layout.addWidget(print_btn)
        export_layout.addStretch()
        
        overtime_layout.addLayout(export_layout)
        
        parent_layout.addWidget(overtime_group)
        
    def setup_attendance_table(self):
        """Setup the attendance report table."""
        headers = [
            "الرقم الوظيفي", "الاسم", "القسم", "الوردية",
            "وقت الحضور", "وقت الانصراف", "الحالة"
        ]
        self.attendance_table.setColumnCount(len(headers))
        self.attendance_table.setHorizontalHeaderLabels(headers)
        self.attendance_table.setAlternatingRowColors(True)
        self.attendance_table.horizontalHeader().setStretchLastSection(True)
        self.attendance_table.verticalHeader().setVisible(False)
        
        # Set column widths
        self.attendance_table.setColumnWidth(0, 120)  # ID
        self.attendance_table.setColumnWidth(1, 200)  # Name
        self.attendance_table.setColumnWidth(2, 150)  # Department
        self.attendance_table.setColumnWidth(3, 120)  # Shift
        self.attendance_table.setColumnWidth(4, 120)  # Check in
        self.attendance_table.setColumnWidth(5, 120)  # Check out
        self.attendance_table.setColumnWidth(6, 120)  # Status
        
    def setup_overtime_table(self):
        """Setup the overtime report table."""
        headers = [
            "الرقم الوظيفي", "الاسم", "القسم", "الوردية",
            "عدد الأيام", "إجمالي الساعات", "متوسط الساعات"
        ]
        self.overtime_table.setColumnCount(len(headers))
        self.overtime_table.setHorizontalHeaderLabels(headers)
        self.overtime_table.setAlternatingRowColors(True)
        self.overtime_table.horizontalHeader().setStretchLastSection(True)
        self.overtime_table.verticalHeader().setVisible(False)
        
        # Set column widths
        self.overtime_table.setColumnWidth(0, 120)  # ID
        self.overtime_table.setColumnWidth(1, 200)  # Name
        self.overtime_table.setColumnWidth(2, 150)  # Department
        self.overtime_table.setColumnWidth(3, 120)  # Shift
        self.overtime_table.setColumnWidth(4, 100)  # Days count
        self.overtime_table.setColumnWidth(5, 120)  # Total hours
        self.overtime_table.setColumnWidth(6, 120)  # Average hours
        
    def load_employees(self):
        """تحميل قائمة الموظفين"""
        employees = self.db.get_all_employees()
        self.emp_combo.clear()
        self.emp_combo.addItem("كل الموظفين", None)
        for emp in employees:
            department = emp.get('department_name') or emp.get('department', '')
            self.emp_combo.addItem(f"{emp['name']} ({department})", emp['id'])

    def load_departments(self):
        """تحميل قائمة الأقسام"""
        departments = self.db.get_all_departments()
        self.dept_combo.clear()
        self.dept_combo.addItem("كل الأقسام", None)
        for dept in departments:
            self.dept_combo.addItem(dept['name'], dept['code'])

    def refresh_reports(self):
        """تحديث كافة التقارير"""
        self.refresh_attendance_report()
        self.refresh_overtime_report()
        
    def refresh_attendance_report(self):
        """Refresh the attendance report table."""
        try:
            date = self.date_edit.date().toString("yyyy-MM-dd")
            dept_code = self.dept_combo.currentData()
            
            # جلب البيانات من قاعدة البيانات
            if dept_code:
                records = self.db.get_department_attendance(dept_code, date)
            else:
                records = self.db.get_all_attendance(date)
                
            # تحديث الجدول
            self.attendance_table.setRowCount(len(records))
            for i, record in enumerate(records):
                # حساب ساعات العمل إذا كان هناك وقت دخول وخروج
                working_hours = ""
                if record['check_in_time'] and record['check_out_time']:
                    check_in = datetime.strptime(record['check_in_time'], "%H:%M:%S")
                    check_out = datetime.strptime(record['check_out_time'], "%H:%M:%S")
                    duration = check_out - check_in
                    working_hours = f"{duration.total_seconds() / 3600:.2f}"
                
                items = [
                    QTableWidgetItem(record['id']),
                    QTableWidgetItem(record['name']),
                    QTableWidgetItem(record['department']),
                    QTableWidgetItem(record['shift_name'] or ""),
                    QTableWidgetItem(record['check_in_time'] or ""),
                    QTableWidgetItem(record['check_out_time'] or ""),
                    QTableWidgetItem(record['status_type'] or "حاضر")
                ]
                
                for j, item in enumerate(items):
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    
                    # تلوين حالة الحضور
                    if j == 6:  # عمود الحالة
                        status = item.text()
                        if status == "حاضر":
                            item.setForeground(QColor("#2b8a3e"))
                        elif status in ["متأخر", "مبكر"]:
                            item.setForeground(QColor("#e67700"))
                        elif status in ["غائب", "إجازة"]:
                            item.setForeground(QColor("#c92a2a"))
                            
                    self.attendance_table.setItem(i, j, item)
                    
        except Exception as e:
            QMessageBox.warning(
                self,
                "خطأ",
                f"حدث خطأ أثناء تحديث تقرير الحضور: {str(e)}",
                QMessageBox.StandardButton.Ok
            )
                
    def refresh_overtime_report(self):
        """Refresh the overtime report table."""
        try:
            from_date = self.from_date.date().toString("yyyy-MM-dd")
            to_date = self.to_date.date().toString("yyyy-MM-dd")
            dept_code = self.overtime_dept_combo.currentData()
            
            # جلب البيانات من قاعدة البيانات
            if dept_code:
                records = self.db.get_department_overtime(dept_code, from_date, to_date)
            else:
                records = self.db.get_all_overtime(from_date, to_date)
                
            # تحديث الجدول
            self.overtime_table.setRowCount(len(records))
            for i, record in enumerate(records):
                items = [
                    QTableWidgetItem(record['id']),
                    QTableWidgetItem(record['name']),
                    QTableWidgetItem(record['department']),
                    QTableWidgetItem(record['shift_name'] or ""),
                    QTableWidgetItem(str(record['days_count'])),
                    QTableWidgetItem(f"{record['total_hours']:.2f}"),
                    QTableWidgetItem(f"{record['average_hours']:.2f}")
                ]
                
                for j, item in enumerate(items):
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    
                    # تلوين ساعات العمل الإضافي
                    if j in [5, 6]:  # أعمدة الساعات
                        hours = float(item.text())
                        if hours > 3:
                            item.setForeground(QColor("#c92a2a"))
                        elif hours > 2:
                            item.setForeground(QColor("#e67700"))
                        else:
                            item.setForeground(QColor("#2b8a3e"))
                            
                    self.overtime_table.setItem(i, j, item)
                    
        except Exception as e:
            QMessageBox.warning(
                self,
                "خطأ",
                f"حدث خطأ أثناء تحديث تقرير الساعات الإضافية: {str(e)}",
                QMessageBox.StandardButton.Ok
            )
                
    def export_report(self, report_type: str, format_type: str):
        """Export report data to Excel or PDF."""
        try:
            # Get the appropriate table
            table = self.attendance_table if report_type == "attendance" else self.overtime_table
            
            # Create DataFrame from table data
            headers = []
            for j in range(table.columnCount()):
                headers.append(table.horizontalHeaderItem(j).text())
            
            data = []
            for i in range(table.rowCount()):
                row = []
                for j in range(table.columnCount()):
                    item = table.item(i, j)
                    # Handle empty items and None values
                    if item and item.text().strip():
                        row.append(item.text())
                    else:
                        row.append('N/A')  # Replace empty values with 'N/A'
                data.append(row)
            
            df = pd.DataFrame(data, columns=headers)
            
            # Get save file path
            file_filter = "Excel Files (*.xlsx)" if format_type == "excel" else "PDF Files (*.pdf)"
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "حفظ التقرير",
                "",
                file_filter
            )
            
            if file_path:
                if format_type == "excel":
                    # Ensure correct file extension
                    if not file_path.endswith('.xlsx'):
                        file_path += '.xlsx'
                    
                    # Create a Pandas Excel writer using openpyxl
                    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                        # Create a new workbook if needed
                        if not hasattr(writer, 'book'):
                            writer.book = Workbook()
                            
                        # Create or select the sheet
                        if 'التقرير' in writer.book.sheetnames:
                            writer.book.remove(writer.book['التقرير'])
                        worksheet = writer.book.create_sheet('التقرير')
                        writer.sheets['التقرير'] = worksheet
                        
                        # Write data to the sheet
                        df.to_excel(
                            writer,
                            index=False,
                            sheet_name='التقرير'
                        )
                        
                        # Adjust column widths
                        worksheet = writer.sheets['التقرير']
                        for idx, col in enumerate(df.columns):
                            max_len = max((
                                df[col].astype(str).str.len().max(),
                                len(str(col)) + 2  # Add padding for Arabic text
                            ))
                            worksheet.column_dimensions[chr(65 + idx)].width = max_len + 2
                            
                        # Remove default empty sheet if exists
                        if 'Sheet' in writer.book.sheetnames:
                            del writer.book['Sheet']
                else:
                    # Convert DataFrame to HTML and save as PDF
                    html = df.to_html(index=False, justify='right')
                    
                    # Add RTL and styling
                    html = f"""
                    <html dir="rtl">
                    <head>
                        <style>
                            body {{ font-family: Arial, sans-serif; }}
                            table {{ width: 100%; border-collapse: collapse; }}
                            th, td {{ 
                                border: 1px solid #ddd; 
                                padding: 8px; 
                                text-align: center; 
                            }}
                            th {{ background-color: #f8f9fa; }}
                        </style>
                    </head>
                    <body>
                        {html}
                    </body>
                    </html>
                    """
                    
                    # Save as PDF
                    printer = QPrinter()
                    printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
                    printer.setOutputFileName(file_path)
                    
                    document = QTextDocument()
                    document.setHtml(html)
                    document.print(printer)
                    
                QMessageBox.information(
                    self,
                    "نجاح",
                    "تم تصدير التقرير بنجاح",
                    QMessageBox.StandardButton.Ok
                )
        except Exception as e:
            QMessageBox.warning(
                self,
                "خطأ",
                f"حدث خطأ أثناء تصدير التقرير: {str(e)}",
                QMessageBox.StandardButton.Ok
            )
                
    def print_report(self, report_type: str):
        """Print the report."""
        # Get the appropriate table
        table = self.attendance_table if report_type == "attendance" else self.overtime_table
        
        printer = QPrinter()
        dialog = QPrintDialog(printer, self)
        
        if dialog.exec() == QPrintDialog.DialogCode.Accepted:
            # Create HTML content
            html = "<html dir='rtl'><body>"
            
            # Add title
            if report_type == "attendance":
                date_str = self.date_edit.date().toString("yyyy-MM-dd")
                title = f"تقرير الحضور اليومي - {date_str}"
            else:
                title = "تقرير الساعات الإضافية"
            html += f"<h2 style='text-align: center'>{title}</h2>"
            
            # Add table
            html += "<table border='1' cellspacing='0' cellpadding='4' style='width: 100%; border-collapse: collapse;'>"
            
            # Add headers
            html += "<tr>"
            for j in range(table.columnCount()):
                header = table.horizontalHeaderItem(j).text()
                html += f"<th style='background-color: #f8f9fa;'>{header}</th>"
            html += "</tr>"
            
            # Add data
            for i in range(table.rowCount()):
                html += "<tr>"
                for j in range(table.columnCount()):
                    item = table.item(i, j)
                    value = item.text() if item else ""
                    html += f"<td style='text-align: center;'>{value}</td>"
                html += "</tr>"
                
            html += "</table></body></html>"
            
            # Create text document and print
            document = QTextDocument()
            document.setHtml(html)
            document.print(printer)

    def load_data(self):
        """تحميل البيانات الأولية للتقارير"""
        self.load_employees()
        self.load_departments()
        self.refresh_reports() 